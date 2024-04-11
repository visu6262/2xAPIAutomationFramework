import openpyxl
# wb=openpyxl.load_workbook("C:\\Users\\91939\\Downloads\\sample3.xlsx")
wb=openpyxl.load_workbook(r"C:\Users\91939\Downloads\sample3.xlsx")
print(type(wb))
sheets=wb.sheetnames
print(sheets) #'Sheet1', 'MyLinks'
# C:\Users\91939\Desktop
sh1=wb["Sheet1"]
print(type(sh1))
print(sh1.cell(1,1).value,end=" ")
print(sh1.cell(1,2).value)
row=sh1.max_row
print(row)
coloum=sh1.max_column
print(coloum)
